import argparse
import os
import csv
import utm
from math import hypot, inf
from openpyxl import load_workbook
from collections import OrderedDict
from datetime import timedelta
from itertools import tee
from dateutil.parser import parse
from copy import deepcopy


def get_console_arguments():
    argument_parser = argparse.ArgumentParser()
    argument_parser.add_argument(
        '-b',
        '--bathymetry_directory',
        default='bathymetry_data/',
        help='Enter path of directory containing '
             '*.csv files with bathymetry data.'
    )
    argument_parser.add_argument(
        '-f',
        '--fairway_points_filepath',
        default='fairway_points.csv',
        help='Enter path of *.csv file containing '
             'points located along a fairway.'
    )
    argument_parser.add_argument(
        '-l',
        '--logger_points_filepath',
        default='logger_points.csv',
        help='Enter path of *.csv file containing '
             'points where water elevation measurements provided.'
    )
    argument_parser.add_argument(
        '-x',
        '--logger_data_filepath',
        default='logger_data.xlsx',
        help='Enter path of *.xlsx file containing water elevation data.'
    )
    argument_parser.add_argument(
        '-o',
        '--output_filepath',
        default='output.csv',
        help='Enter path of *.csv file containing the script`s output.'
    )
    arguments = argument_parser.parse_args()
    return arguments


def get_bathymetry_file_paths(directory_path):
    filename_list = []
    for entry in os.scandir(directory_path):
        file_root, file_extension = os.path.splitext(entry.path)
        if file_extension == '.csv':
            filename_list.append(entry.path)
    return filename_list


def load_csv_data(file_name_list):
    data_list = []
    try:
        for file_path in file_name_list:
            with open(file_path, 'r', encoding='utf-8') as input_file:
                file_reader = csv.reader(input_file, delimiter=';')
                for row in file_reader:
                    row.append(file_path)
                    data_list.append(row)
        return data_list
    except FileNotFoundError:
        return None


def load_input_data(csv_file_names, xlsx_file_name):
    input_files_content = []
    for file_type in csv_file_names:
        filename_list = csv_file_names[file_type]
        data_of_single_input_type = load_csv_data(filename_list)
        input_files_content.append(data_of_single_input_type)
    try:
        xlsx_workbook = load_workbook(xlsx_file_name, read_only=True)
    except FileNotFoundError:
        xlsx_workbook = None
    input_files_content.append(xlsx_workbook)
    return input_files_content


def get_bathymetry_points(point_list):
    bathymetry_list = []
    for point in point_list:
        feature_names = ['longitude', 'latitude', 'depth', 'time', 'filepath']
        try:
            long, lat, depth, _, _, _, time, filepath = point
        except ValueError:
            return None
        try:
            depth = float(depth.replace(',', '.'))
        except ValueError:
            depth = None
        feature_values = [long, lat, depth, time, filepath]
        verified_feature_values = []
        for feature in feature_values:
            if not feature:
                feature = None
            verified_feature_values.append(feature)
        point_dict = dict(zip(feature_names, verified_feature_values))
        bathymetry_list.append(point_dict)
    return bathymetry_list


def get_fairway_points(point_list):
    fairway_list = []
    for point in point_list:
        feature_names = [
            'longitude',
            'latitude',
            'distance_from_seashore',
            'file_name'
        ]
        try:
            longitude, latitude, _, distance, file_name = point
        except ValueError:
            return None
        feature_values = [longitude, latitude, distance, file_name]
        point_dict = dict(zip(feature_names, feature_values))
        fairway_list.append(point_dict)
    return fairway_list


def get_logger_points(point_list):
    logger_list = []
    for point in point_list:
        feature_names = ['longitude', 'latitude', 'logger_name', 'file_name']
        try:
            longitude, latitude, logger_name, file_name = point
        except ValueError:
            return None
        feature_values = [longitude, latitude, logger_name, file_name]
        point_dict = dict(zip(feature_names, feature_values))
        logger_list.append(point_dict)
    return logger_list


def get_logger_data(xlsx_workbook):
    all_loggers_data = {}
    for sheet in xlsx_workbook:
        logger_trace = {}
        for row in sheet.iter_rows(min_row=2, max_col=2):
            measurement_datetime = row[0].value
            if measurement_datetime is None:
                break
            elevation = row[1].value
            logger_trace[measurement_datetime] = elevation
        all_loggers_data[sheet.title] = logger_trace
    return all_loggers_data


def convert_geocoordinates_to_utm(dataset_list):
    dataset_with_coordinates = deepcopy(dataset_list)
    for dataset in dataset_with_coordinates:
        for point in dataset:
            if not all(point.values()):
                point['longitude'] = None
                point['latitude'] = None
                continue
            longitude = float(point['longitude'])
            latitude = float(point['latitude'])
            utm_long, utm_lat, zone_num, zone_letter = utm.from_latlon(
                latitude, longitude
            )
            point['longitude'] = utm_long
            point['latitude'] = utm_lat
    return dataset_with_coordinates


def round_logger_datetime(logger_data):
    data_with_rounded_time = deepcopy(logger_data)
    for logger in logger_data:
        logger_trace = logger_data[logger]
        rounded_logger_data = {}
        for measurement_datetime, water_elevation in logger_trace.items():
            measurement_datetime += timedelta(seconds=30)
            rounded_datetime = measurement_datetime - timedelta(
                seconds=measurement_datetime.second,
                microseconds=measurement_datetime.microsecond
            )
            rounded_logger_data[rounded_datetime] = water_elevation
        data_with_rounded_time[logger] = rounded_logger_data
    return data_with_rounded_time


def get_distance_to_the_fairway_point(fairway_point, lat, long):
    fairway_point_lat = fairway_point['latitude']
    fairway_point_long = fairway_point['longitude']
    only_numbers = fairway_point_lat and fairway_point_long and lat and long
    if only_numbers:
        distance = hypot(fairway_point_lat - lat, fairway_point_long - long)
        return distance
    return inf


def get_distance_from_sea(points, points_along_fairway):
    points_with_distance_from_sea = deepcopy(points)
    for point in points_with_distance_from_sea:
        if not all(point.values()):
            point['distance_from_seashore'] = None
            continue
        latitude = point['latitude']
        longitude = point['longitude']
        closest_fairway_point = min(
            points_along_fairway,
            key=lambda x: get_distance_to_the_fairway_point(
                x,
                latitude,
                longitude
            )
        )
        distance_from_sea = closest_fairway_point['distance_from_seashore']
        point['distance_from_seashore'] = float(distance_from_sea)
    return points_with_distance_from_sea


def get_nearest_loggers(distance_from_seashore, logger_list):
    logger_list.sort(key=lambda x: x['distance_from_seashore'])
    logger_iterator, logger_iterator_duplicate = tee(logger_list)
    next(logger_iterator_duplicate)
    pairwise_logger_list = zip(logger_iterator, logger_iterator_duplicate)

    for lower_logger, upper_logger in pairwise_logger_list:
        low_log_distance = lower_logger['distance_from_seashore']
        up_log_distance = upper_logger['distance_from_seashore']
        if distance_from_seashore < low_log_distance:
            return lower_logger, upper_logger
        if low_log_distance <= distance_from_seashore < up_log_distance:
            return lower_logger, upper_logger

    uppermost_loggers = (lower_logger, upper_logger)
    return uppermost_loggers


def interpolate_water_surface(
        lower_level,
        upper_level,
        lower_distance,
        upper_distance,
        measurement_distance
):
    water_slope = (lower_level-upper_level)/(lower_distance-upper_distance)
    y_intercept = upper_level - water_slope*upper_distance
    water_elevation = water_slope * measurement_distance + y_intercept
    return water_elevation


def get_loggers_working_at_measurement_time(
        list_of_logger_points,
        logger_data,
        measurement_datetime
):
    not_working_logger_names = []
    list_of_working_loggers = deepcopy(list_of_logger_points)
    for logger_name in logger_data:
        if measurement_datetime not in logger_data[logger_name]:
            not_working_logger_names.append(logger_name)
    for logger_point in list_of_working_loggers:
        if logger_point['logger_name'] in not_working_logger_names:
            list_of_working_loggers.remove(logger_point)
    return list_of_working_loggers, not_working_logger_names


def get_water_elevation(bathymetry, logger_data_points, logger_traces):
    disabled_logs = []
    bathymetry_points = deepcopy(bathymetry)
    for measurement_point in bathymetry_points:
        if not all(measurement_point.values()):
            measurement_point['water_elevation'] = None
            continue
        measurement_time = parse(
            measurement_point['time'],
            dayfirst=True,
            yearfirst=False
        )
        working_logs, disabled_logs = get_loggers_working_at_measurement_time(
            logger_data_points,
            logger_traces,
            measurement_time
        )
        lower_log, upper_log = get_nearest_loggers(
            measurement_point['distance_from_seashore'],
            working_logs
        )
        lower_log_name = lower_log['logger_name']
        upper_log_name = upper_log['logger_name']
        lower_elevation = logger_traces[lower_log_name][measurement_time]
        upper_elevation = logger_traces[upper_log_name][measurement_time]
        water_elevation = interpolate_water_surface(
            lower_elevation,
            upper_elevation,
            lower_log['distance_from_seashore'],
            upper_log['distance_from_seashore'],
            measurement_point['distance_from_seashore']
        )
        measurement_point['water_elevation'] = water_elevation
    return bathymetry_points, disabled_logs


def get_bottom_elevation(bathymetry):
    bathymetry_with_bottom_elevation = deepcopy(bathymetry)
    for point in bathymetry_with_bottom_elevation:
        if not all(point.values()):
            point['bottom_elevation'] = None
            continue
        water_elevation = point['water_elevation']
        depth = point['depth']
        bottom_elevation = water_elevation - depth
        point['bottom_elevation'] = bottom_elevation
    return bathymetry_with_bottom_elevation


def print_about_filenotfounderror_and_exit(
        input_content,
        csv_filenames,
        xlsx_filename
):
    bathymetry, fairway, logger_data, water_elevation = input_content
    if bathymetry is None:
        exit('Can not find {}'.format(csv_filenames['bathymetry']))
    if fairway is None:
        exit('Can not find {}'.format(csv_filenames['points_along_fairway']))
    if logger_data is None:
        exit('Can not find {}'.format(csv_filenames['logger_coordinates']))
    if water_elevation is None:
        exit('Can not find {}'.format(xlsx_filename))
    return


def print_about_wrong_file_format_and_exit(
        input,
        csv_filenames
):
    bathymetry, fairway_info, logger_info, water_elevation_info = input
    if bathymetry is None:
        exit(
            'The wrong format of data in some of these files: {}.'.format(
                csv_filenames['bathymetry']
            )
        )
    if fairway_info is None:
        exit(
            'The wrong format of data in {}'.format(
                csv_filenames['points_along_fairway']
            )
        )
    if logger_info is None:
        exit(
            'The wrong format of data in {}'.format(
                csv_filenames['logger_coordinates']
            )
        )
    if water_elevation_info is None:
        exit('The wrong format of *.xlsx file.')
    return


def print_invalid_points(points):
    for point in points:
        if not all(point.values()):
            print('WARNING! Invalid point: {}'.format(point))
    return


def output_result(bathymetry_info, output_path):
    field_names = [
        'longitude',
        'latitude',
        'bottom_elevation',
        'time',
        'water_elevation',
        'depth',
        'distance_from_seashore',
        'filepath'
    ]
    with open(output_path, 'w', newline='', encoding='utf-8') as output_file:
        writer = csv.DictWriter(
            output_file,
            fieldnames=field_names,
            delimiter=';'
        )
        writer.writeheader()
        for point in bathymetry_info:
            writer.writerow(point)


def get_input_filenames(script_arguments):
    bathymetry_file_paths_list = get_bathymetry_file_paths(
        script_arguments.bathymetry_directory
    )
    # use OrderedDict() instance to correctly extract data
    # from output of "load_input_data()"
    input_csv_filenames = OrderedDict([
        ('bathymetry', bathymetry_file_paths_list),
        ('points_along_fairway', [script_arguments.fairway_points_filepath]),
        ('logger_coordinates', [script_arguments.logger_points_filepath])
    ])
    water_elevation_filename = script_arguments.logger_data_filepath
    return input_csv_filenames, water_elevation_filename


def get_data_from_files_content(content):
    bathymetry_data, fairway_data, loggers, xlsx_file_workbook = content
    bathymetry_points = get_bathymetry_points(bathymetry_data)
    fairway_points = get_fairway_points(fairway_data)
    logger_points = get_logger_points(loggers)
    water_elevation_data = get_logger_data(xlsx_file_workbook)
    input_data = [
        bathymetry_points,
        fairway_points,
        logger_points,
        water_elevation_data
    ]
    return input_data


if __name__ == "__main__":
    console_arguments = get_console_arguments()
    csv_filenames, xlsx_filename = get_input_filenames(console_arguments)
    input_files_content = load_input_data(csv_filenames, xlsx_filename)
    if None in input_files_content:
        print_about_filenotfounderror_and_exit(
            input_files_content,
            csv_filenames,
            xlsx_filename
        )

    input_data = get_data_from_files_content(input_files_content)
    if None in input_data:
        print_about_wrong_file_format_and_exit(
            input_data,
            csv_filenames,
        )
    bathymetry_points, fairway_points, logger_points, water_elevation_data = input_data
    datasets = [bathymetry_points, fairway_points, logger_points]
    utm_bathymetry, utm_fairway, utm_loggers = convert_geocoordinates_to_utm(
        datasets
    )
    bathymetry_points = get_distance_from_sea(utm_bathymetry, utm_fairway)
    logger_points = get_distance_from_sea(utm_loggers, utm_fairway)
    water_elevation_data = round_logger_datetime(water_elevation_data)
    bathymetry_points, switched_off_loggers = get_water_elevation(
        bathymetry_points,
        logger_points,
        water_elevation_data
    )
    if switched_off_loggers:
        print(
            'WARNING! These loggers have no data for some '
            'measurement time moments: {}'.format(switched_off_loggers)
        )
    bathymetry_points_with_bottom_elevation = get_bottom_elevation(
        bathymetry_points
    )
    print_invalid_points(bathymetry_points_with_bottom_elevation)
    output_result(
        bathymetry_points_with_bottom_elevation,
        console_arguments.output_filepath
    )
